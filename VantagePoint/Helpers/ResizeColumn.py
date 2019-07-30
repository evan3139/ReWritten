import openpyxl
from openpyxl.utils import get_column_letter


def resize_columns(excel_name):
    # This reopens the excel file but in the openpyxl library allowing us to alter column lengths
    wb = openpyxl.load_workbook(excel_name)
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the Column Name Here
            for cell in col:
                try:  # Needed to avoid empty cell errors
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(excel_name)