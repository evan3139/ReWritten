import openpyxl
import statistics
import pandas as pd
import numpy as np


def intake_averages(excel, groups):
    wb = openpyxl.load_workbook(excel)

    for sheet in wb.worksheets:
        # Since we cant update the max of the sheet after changing it.
        # We grab it at the beginning of every sheet and set it to a variable
        # This matter for putting the averages past the bottom row
        max_row = sheet.max_row
        for col in sheet.columns:
            if col[0].value is not None and col[0].value.lower() in groups:
                average = 0
                total = 0
                deviation = []
                # This depicts where the average will be placed. So for every 3 Nones,
                # The middle one will get the average placed.
                row_tick = 1
                for index, cell in enumerate(col):
                    if cell.value is not None and not isinstance(cell.value, str):
                        deviation.append(cell.value)
                        average += cell.value
                        total += 1
                    elif cell.value is None:
                        row_tick += 1
                    if row_tick == 3:
                        try:
                            cell.value = average / total
                        except:
                            cell.value = 0
                        # Stop it from overwiting the other averages.
                        if sheet.cell(row=index + 1, column=cell.col_idx - 1).value is None:
                            sheet.cell(row=index + 1, column=cell.col_idx - 1).value = "Avg"

                        average = total = 0
                    elif row_tick == 4:
                        try:
                            cell.value = statistics.stdev(deviation)
                        except:
                            cell.value = 0

                        if sheet.cell(row=index + 1, column=cell.col_idx - 1).value is None:
                            sheet.cell(row=index + 1, column=cell.col_idx - 1).value = "STDEV"

                        row_tick = 0
                        deviation = []
                    # Special Case since it wont get any None For the end of the col.
                    # We must find when the for loop is about to end.
                    # I hate that this is here, so if you can figure something out plz fix.
                    if index + 1 == max_row:
                        if sheet.cell(row=max_row + 2, column=cell.col_idx - 1).value is None and sheet.cell(
                                row=max_row + 3, column=cell.col_idx - 1).value is None:
                            sheet.cell(row=max_row + 2, column=cell.col_idx - 1).value = "Avg"
                            sheet.cell(row=max_row + 3, column=cell.col_idx - 1).value = "STDEV"
                        sheet.cell(row=max_row + 2, column=cell.col_idx).value = average / total
                        sheet.cell(row=max_row + 3, column=cell.col_idx).value = statistics.stdev(deviation)

    wb.save(excel)
