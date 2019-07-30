import openpyxl
from openpyxl import *
from Helpers.ReadDocument import read_content, read_header


def create_intake_header(file, excel, header_list):
    headers = read_header(file)
    counter = 1
    score_counter = 1

    wb = load_workbook(excel)
    sheets = wb.get_sheet_names()

    # Renamed these just for clarity. Couldve done sheets[0] or sheets [1]
    full = wb[sheets[0]]
    scores = wb[sheets[1]]

    for header in headers:
        full.cell(row=1, column=counter).value = header
        counter += 1
        if header.lower() in header_list:
            scores.cell(row=1,column=score_counter).value = header
            score_counter += 1

    wb.save(excel)
