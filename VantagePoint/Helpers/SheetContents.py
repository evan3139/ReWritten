from Helpers.ReadDocument import read_content, read_header
import openpyxl
from openpyxl import *


def fill_intake_sheet(file,excel,headers):
    everything = read_content(file)

    # Get all the sheets in the excel
    wb = load_workbook(excel)
    sheets = wb.get_sheet_names()

    # Same thing as before with making the sheets easier to see in the code
    full = wb[sheets[0]]
    scores = wb[sheets[1]]

    # Get the max row and save it as a variable to save on complexity.
    # Row will be the same for both sheets, The Cols will not though
    max_row = full.max_row + 1
    full_col = 1
    score_col = 1

    for content in everything:
        if content[1].isdigit() or content[1].isdecimal():
            full.cell(row=max_row, column=full_col).value = int(content[1])
            full_col += 1
        else:
            full.cell(row=max_row, column=full_col).value = content[1]
            full_col += 1
        if content[0].lower() in headers:
            if content[1].isdigit() or content[1].isdecimal():
                scores.cell(row=max_row, column=score_col).value = int(content[1])
                score_col += 1
            else:
                scores.cell(row=max_row, column=score_col).value = content[1]
                score_col += 1

    wb.save(excel)